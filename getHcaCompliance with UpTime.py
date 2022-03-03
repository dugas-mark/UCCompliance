
from paramiko.ssh_exception import AuthenticationException
from requests import Session
from requests.auth import HTTPBasicAuth
from zeep import Client, Settings
from zeep.exceptions import Fault
from zeep.transports import Transport
import paramiko
import urllib3
from datetime import datetime, timedelta
from openpyxl.styles import Alignment, Color, Font

def getRegisteredDeviceCnt(username, passwd, cucm_ip):
	"""
	Verified no media resources are unregistered
	"""
	urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
	session = Session()
	session.auth = HTTPBasicAuth(username, passwd)
	session.verify = False
	transport = Transport(session=session)
	# Parameters to conduct search
	items = {'item': [{'Item': '*'}]}
	client = Client(f'https://{cucm_ip}:8443/realtimeservice2/services/RISService70?wsdl', transport=transport)
	criteria_type = client.get_type('ns0:CmSelectionCriteria')


	service = client.create_service('{http://schemas.cisco.com/ast/soap}RisBinding', 
									f'https://{cucm_ip}:8443/realtimeservice2/services/RISService70')
	regDeviceCnt = {'Registered':{'Phone':'','Cti':''},'UnRegistered':{'Phone':'','Cti':''}}

	for r in regDeviceCnt:
		for t in regDeviceCnt[r]:
			criteria = criteria_type(MaxReturnedDevices=1000, DeviceClass=t, Model=255, Status=r,
							SelectBy='Name', Protocol='Any', DownloadStatus='Any',SelectItems=items)
			resp = service.selectCmDevice(CmSelectionCriteria=criteria,StateInfo='')
			regDeviceCnt[r][t] = resp['SelectCmDeviceResult']['TotalDevicesFound'] 

	return regDeviceCnt

def mrglCheck(username, passwd, cucm_ip):

#	Compliance script mrglCheck
#	michael hagans, DOF9318
#	8/6/2020
#
#   Verified no media resources are unregistered

	session = Session()
	session.auth = HTTPBasicAuth(username, passwd)
	session.verify = False
	transport = Transport(session=session)
	items = {'item': [{'Item': '*'}]}
	# Parameters to conduct search

	client = Client(
		f'https://{cucm_ip}:8443/realtimeservice2/services/RISService70?wsdl', transport=transport)
	criteria_type = client.get_type('ns0:CmSelectionCriteria')
	criteria = criteria_type(MaxReturnedDevices=1000, DeviceClass='MediaResources', Model=255, Status='UnRegistered',
							SelectBy='Name', Protocol='Any', DownloadStatus='Any', SelectItems=items)

	service = client.create_service('{http://schemas.cisco.com/ast/soap}RisBinding',
									f'https://{cucm_ip}:8443/realtimeservice2/services/RISService70')

	resp = service.selectCmDevice(CmSelectionCriteria=criteria, StateInfo='')
	if resp['SelectCmDeviceResult']['TotalDevicesFound'] != 0:
		returnMsg = 'Media Resources Found Unregistered. Verify MR Status.'
	else:
		returnMsg = 'All Media Resources Registered.'

	return returnMsg


def getBackupStatus(username, passwd, cucm_ip, timeout=60):
	import re

	session = Session()
	session.verify = False
	session.auth = HTTPBasicAuth(username, passwd)
	transport = Transport(session=session, timeout=timeout,
						  operation_timeout=timeout)
	settings = Settings(strict=False, xml_huge_tree=True)
	client = Client(
		f'https://{cucm_ip}:8443/platform-services/services/MaintenanceService?wsdl', transport=transport, settings=settings)
	service = client.create_service('{http://services.api.platform.vos.cisco.com}MaintenanceServiceSoap12Binding',
									f'https://{cucm_ip}:8443/platform-services/services/MaintenanceService.MaintenanceServiceHttpsSoap12Endpoint')

	result = service.getBackupProgress()
	if result.backupProgressResult['status'].__contains__('SUCCESS'):
		backupStatus = 'Success'
		#d = re.search('(Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s\d\d\s\d\d:\d\d:\d\d\s\w\w\w\s\d\d\d\d',
		#					result.backupProgressResult['componentList'])
		#dateBkup = d.string[d.regs[0][0]: d.regs[0][1]]
		# Use the backup file name to retrieve date of backup
		d = re.search('\d\d\d\d-\d\d-\d\d',result.backupProgressResult['tarFile'])
		dateBkup = d.group()
	else:
		backupStatus = 'Failed'
		dateBkup = ''

	return backupStatus + ' - ' + dateBkup


def checkStandardPwd(standardOsUserName, standardCcmOsPwd, 
					standardCcmAdminUserName,standardCcmAdminPwd, 
					wsdl_url, cucm_ip, username, passwd):

	# Check CCMADmin OS password via CLI
	standardOsPwd = False
	standardGuiPwd = False
	ssh_client = paramiko.SSHClient()
	ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	try:
		ssh_client.connect(hostname=cucm_ip, username= standardOsUserName, password=standardCcmOsPwd, timeout=60)
		standardOsPwd = True
	except AuthenticationException as e:
		print('\n	Checking standard OS pwd auth failed: ',e,standardOsUserName,' ',standardCcmOsPwd)
		standardOsPwd = False
	except Exception as e:
		print('\n	Checking standard OS pwd other exception: ',e,standardOsUserName,' ',standardCcmOsPwd)
		standardOsPwd = False
	ssh_client.close()

	# standardGuiPwd = requests.get(f'https://{cucm_ip}/ccmadmin/j_security_check', verify=False, auth=HTTPBasicAuth('CCMAdmin',standardCcmAdminPwd))
	session = Session()
	session.verify = False
	session.auth = HTTPBasicAuth(username, passwd)
	transport = Transport(session=session, timeout=10)
	settings = Settings(strict=False, xml_huge_tree=True)
	client = Client(wsdl_url+'AXLAPI.wsdl',
					settings=settings, transport=transport)
	service = client.create_service("{http://www.cisco.com/AXLAPIService/}AXLAPIBinding",
									 f'https://{cucm_ip}:8443/axl/')
	try:
		standardGuiPwd = service.doAuthenticateUser(userid= standardCcmAdminUserName, 
													password=standardCcmAdminPwd)['return']['userAuthenticated'] == 'true'
	except Exception as e:
		# if doAuthenticateUser fails due to "unknown result occurred" mark as password failed
		standardGuiPwd = False

	return {'OS':str(standardOsPwd), 'GUI':str(standardGuiPwd)}

def getListPhones(username, passwd, cucm_ip, wsdl_url):
	session = Session()
	session.verify = False
	session.auth = HTTPBasicAuth(username, passwd)
	transport = Transport(session=session, timeout=10)
	settings = Settings(strict=False, xml_huge_tree=True)

	client = Client(wsdl_url+'AXLAPI.wsdl',
					settings=settings, transport=transport)
	service = client.create_service('{http://www.cisco.com/AXLAPIService/}AXLAPIBinding',
									f'https://{cucm_ip}:8443/axl/')
	response = service.listPhone({'name':'%'},{'name':'','class':'Cti'})
	return response

def getCUCMVer(username, passwd, cucm_ip, wsdl_url):

	sess = Session()
	sess.verify = False
	sess.auth = HTTPBasicAuth(username, passwd)
	trans = Transport(session=sess, timeout=10)
	sett = Settings(strict=False, xml_huge_tree=True)

	clnt = Client(wsdl_url+'AXLAPI.wsdl',
					settings=sett, transport=trans)
	serv = clnt.create_service('{http://www.cisco.com/AXLAPIService/}AXLAPIBinding',
									f'https://{cucm_ip}:8443/axl/')
	response = serv.getCCMVersion()
	return response['return']['componentVersion']['version']

def getCUCMPubInfo(username, passwd, cucm_ip):

	pubInfo = {
				'CallManagerActive':'',
				'CallManagerUpTime':'',
				'CiscoAXLWebSvcTime':''
				}

	session = Session()
	session.verify = False
	session.auth = HTTPBasicAuth(username, passwd)
	transport = Transport(session=session, timeout=10)
	settings = Settings(strict=False, xml_huge_tree=True)
	client = Client(f'https://{cucm_ip}:8443/controlcenterservice2/services/ControlCenterServices?wsdl',
					settings=settings, transport=transport)
	service = client.create_service('{http://schemas.cisco.com/ast/soap}ControlCenterServicesBinding',
									f'https://{cucm_ip}:8443/controlcenterservice2/services/ControlCenterServices')
	# method assumes that server URL is that of publisher; returns ServiceName='Cisco CallManager', ServiceStatus='whatever'
	statusPub = service.soapGetServiceStatus(ServiceStatus=['Cisco CallManager','Cisco AXL Web Service'])
	days = lambda d: 0 if d < 0 else d
	x = timedelta(seconds=statusPub['ServiceInfoList']['item'][0]['UpTime'])
	y = timedelta(seconds=statusPub['ServiceInfoList']['item'][1]['UpTime'])
	pubInfo['CallManagerUpTime'] = statusPub['ServiceInfoList']['item'][0]['ServiceName']+' - Days: '+str(days(x.days))+' Hours: '+str(x.seconds//3600)
	pubInfo['CiscoAXLWebSvcTime'] = statusPub['ServiceInfoList']['item'][1]['ServiceName']+' - Days: '+str(days(y.days))+' Hours: '+str(y.seconds//3600)
	if statusPub['ServiceInfoList']['item'][0]['ServiceStatus'] == 'Started':
		pubInfo['CallManagerActive'] = 'Non-standard:  CM service is running on pub'
	else:
		pubInfo['CallManagerActive'] = 'Standard:  CM service is inactive on pub'

	return pubCallManagerActive


def getCUCMLdap(username, passwd, cucm_ip, wsdl_url):

	session = Session()
	session.verify = False
	session.auth = HTTPBasicAuth(username, passwd)
	transport = Transport(session=session, timeout=10)
	settings = Settings(strict=False, xml_huge_tree=True,
						xsd_ignore_sequence_order=True)
	# ldAccessGrpName = []
	client = Client(wsdl_url+'AXLAPI.wsdl',
					settings=settings, transport=transport)
	service = client.create_service("{http://www.cisco.com/AXLAPIService/}AXLAPIBinding",
									f'https://{cucm_ip}:8443/axl/')

	
	# Check that LDAP is enabled

	#ldapEnabled = service.getLdapSystem()
	if service.getLdapSystem()['return']['ldapSystem']['syncEnabled'] == 'false':
		# LDAP is not enabled on this CUCM, end LDAP check
		return 'LDAP not enabled'

	# LDAP is enabled; check for standard admin access in at least one LDAP Directory
	# Search for conditions:
	#       LDAP User Search Base = 'DC=hca,DC=corpad,DC=net' AND
	#       LDAP Custom Filter for Users contains 'UCE' and 'CDE'
	# If true, LDAP check passes

	# Find all LDAP Directories configured for this cluster.

	nameLdapDirectory = service.listLdapDirectory(
		searchCriteria={'name': '%'}, returnedTags={'name': '', 'userSearchBase':''})
	if nameLdapDirectory['return'] == None:
		# LDAP was set to "on" but there are no LDAP directories configured, end LDAP check
		return 'LDAP directory not configured'
	
	for n in nameLdapDirectory['return']['ldapDirectory']:
		directoryFilter = service.getLdapDirectory(name = n.name,
										returnedTags = {'ldapFilter':''})
		if directoryFilter['return']['ldapDirectory']['ldapFilter']['_value_1'] == None:
			break
		ldapFilters = service.getLdapFilter(name=directoryFilter['return']['ldapDirectory']['ldapFilter']['_value_1'],
											returnedTags = {'name':'','filter':''})

		if (('UCE' in ldapFilters['return']['ldapFilter']['filter']) \
				or ('CDE' in ldapFilters['return']['ldapFilter']['filter'])) \
				and n.userSearchBase == 'DC=hca,DC=corpad,DC=net':
			return 'Standard LDAP'
		else:
			pass

	return 'Non-Standard LDAP'

def getCUCMClusterID(username, passwd, cucm_ip, wsdl_url):
	import re

	session = Session()
	session.verify = False
	session.auth = HTTPBasicAuth(username, passwd)
	transport = Transport(session=session, timeout=10)
	settings = Settings(strict=False, xml_huge_tree=True)

	client = Client(wsdl_url+'AXLAPI.wsdl',
					settings=settings, transport=transport)
	service = client.create_service("{http://www.cisco.com/AXLAPIService/}AXLAPIBinding",
									f'https://{cucm_ip}:8443/axl/')
	response = service.getServiceParameter(
		processNodeName='EnterpriseWideData', name='ClusterID', service='Enterprise Wide')

	clusterIDName = response['return']['serviceParameter']['value']

	# print('\n',re.match('[A-Z][A-Z][A-Z]-[1-9]',clusterIDName))
	
	if re.match('[A-Z][A-Z][A-Z]-[1-9]',clusterIDName) != None:
		return clusterIDName + ' : Standard'
	else:
		return clusterIDName + ' : Non-Standard'

def getCUCMCmgLoad(username, passwd, cucm_ip, wsdl_url):

	session = Session()
	session.verify = False
	session.auth = HTTPBasicAuth(username, passwd)
	transport = Transport(session=session, timeout=10)
	settings = Settings(strict=False, xml_huge_tree=True)

	client = Client(wsdl_url+'AXLAPI.wsdl',
					settings=settings, transport=transport)
	service = client.create_service('{http://www.cisco.com/AXLAPIService/}AXLAPIBinding',
									f'https://{cucm_ip}:8443/axl/')

	deviceQuery = service.executeSQLQuery(sql="""SELECT cmg.name AS CMG_Group, COUNT(d.name) AS IP_Phones 
					FROM device AS d 
					INNER JOIN devicepool AS dp ON d.fkDevicePool=dp.pkid 
					INNER JOIN callmanagergroup AS cmg ON dp.fkcallmanagergroup=cmg.pkid 
					INNER JOIN typemodel AS tm ON tm.enum=d.tkmodel 
					WHERE (tm.name != 'Analog Phone' AND tm.name != 'Conference Bridge'
						AND tm.name != 'CTI Route Point' AND tm.name != 'CTI Port'
						AND tm.name != 'MGCP Station' AND tm.name != 'Route List'
						AND tm.name != 'H.323 Gateway'
						AND tm.name != 'Music on Hold' 
						AND tm.name != 'Media Termination Point' 
						AND tm.name != 'Tone Announcement Player'
						AND tm.name != 'Cisco IOS Conference Bridge (HDV2)'
						AND tm.name != 'Cisco IOS Software Media Termination Point (HDV2)' 
						AND tm.name != 'Cisco IOS Media Termination Point (HDV2)' 
						AND tm.name != 'SIP Trunk' 
						AND (dp.name LIKE '%PH%' OR dp.name LIKE '%iMobile%'))
					GROUP BY cmg.name
					ORDER BY cmg.name""")

	if deviceQuery['return'] == None:
		return 'Unavailable'
	else:
		countByCmg = {}
		for cbdp in deviceQuery['return']['row']:
			countByCmg[cbdp[0].text] = int(cbdp[1].text)
	return countByCmg

def getCUCMOvaInfo(username, passwd, cucm_ip):
	import math

	session = Session()
	session.verify = False
	session.auth = HTTPBasicAuth(username, passwd)
	transport = Transport(session=session, timeout=10)
	settings = Settings(strict=False, xml_huge_tree=True)
	client = Client(f'https://{cucm_ip}:8443/perfmonservice2/services/PerfmonService?wsdl', 
					settings=settings, transport=transport)
	service = client.create_service('{http://schemas.cisco.com/ast/soap}PerfmonBinding',
									f'https://{cucm_ip}:8443/perfmonservice2/services/PerfmonService')
	pfmSession = service.perfmonOpenSession( )

	try:
		responseCpu = service.perfmonListInstance(Host=cucm_ip,Object ='Processor')
		ovaCpu = len(responseCpu) - 1
	except Fault as err:
		print( f'Zeep error: perfmonListInstance: {err}' )

	try:
		responseMem = service.perfmonCollectCounterData(Host =cucm_ip,Object = 'Memory')
		responseMemLen = len(responseMem) -1
		n = 0
		while n <= responseMemLen:
			if responseMem[n]['Name']['_value_1'].__contains__('Total KBytes'):
				ovaMem = int(math.ceil(float(responseMem[n]['Value'])/1024000))
				break
			n += 1
	except Fault as err:
		print( f'Zeep error: perfmonCollectCounterData: {err}' )

	if ovaCpu == 4 and ovaMem == 8:
		return 'Large'
	else:
		return 'non-Standard'


def initWorkSheet(wb,ws):
	#Set up Excel worksheet for this month's Compliance Check
	runDate = datetime.now()
	headers = ['Cluster Name', 'Pub IP', 'CUCM Ver', 'LDAP Status', 'Cluster ID', 'Server Specs', 'Publisher CM Inactive', 'MRGL Registered', 'CMG Load', 'Backup Status', 'Password Status', 'Notes']
	ws.title = f'UC Compliance Check - {runDate.strftime("%b")} {runDate.year}'
	ws.append(headers)
	wb.save(f'UC Compliance {runDate.strftime("%b")} {str(runDate.year)}.xlsx')
	return f'UC Compliance {runDate.strftime("%b")} {str(runDate.year)}.xlsx'
	
def addCell(ws, spreadSheetRow, column, data):
	cellData = ''
	if isinstance(data, str):
		cellData = data
	elif isinstance(data, list) or isinstance(data, dict):
		# Data is multi-line, set wrap text attribute for this cell
		ws.cell(row=spreadSheetRow,column=column).alignment = Alignment(wrapText=True)
		for returned in data:
			cellData += f'{returned}:{data[returned]}\n'
	ws.cell(row=spreadSheetRow,column=column).value=cellData
	#print('\n',ws.cell(row=spreadSheetRow,column=column).value)
	return
